"""
PersonalViewSet — CRUD + DNI lookup + Excel export for RRHH personnel records.
"""
import io

from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import OpenApiParameter, extend_schema, extend_schema_view
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.core.models import Personal
from apps.core.permissions import IsHRManager, IsPasajesStaff
from apps.core.serializers.personal import (
    PersonalDetailSerializer,
    PersonalDNILookupSerializer,
    PersonalListSerializer,
)


@extend_schema_view(
    list=extend_schema(tags=['personal'], summary='List personnel'),
    retrieve=extend_schema(tags=['personal'], summary='Get personnel detail'),
    create=extend_schema(tags=['personal'], summary='Create personnel record'),
    update=extend_schema(tags=['personal'], summary='Update personnel record'),
    partial_update=extend_schema(tags=['personal'], summary='Partial update personnel record'),
    destroy=extend_schema(tags=['personal'], summary='Delete personnel record'),
)
class PersonalViewSet(viewsets.ModelViewSet):
    """
    Full CRUD for Personal (employee) records.

    SYSPCC-006 FIX 1: `Personal` holds sensitive RRHH data (salary, banking,
    address, emergency contact, family). `list`/`retrieve`/`export` are
    restricted to RRHH-privileged roles (IsHRManager). `buscar_dni` stays
    available to the pasajes module (IsPasajesStaff) but returns only
    non-sensitive fields via PersonalDNILookupSerializer. Write operations
    (create/update/delete) still require staff privileges.
    """

    queryset = Personal.objects.select_related('proyecto', 'user').all()
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['estado', 'proyecto', 'sede', 'guardia', 'condicion_trabajo']
    search_fields = ['dni', 'apellidos_nombres', 'puesto', 'celular']
    ordering_fields = ['apellidos_nombres', 'fecha_ingreso', 'estado', 'puesto']
    ordering = ['apellidos_nombres']

    def get_serializer_class(self):
        if self.action == 'list':
            return PersonalListSerializer
        if self.action == 'buscar_dni':
            return PersonalDNILookupSerializer
        return PersonalDetailSerializer

    def get_permissions(self):
        """
        SYSPCC-006 FIX 1:
        - `buscar_dni` (pasajes autofill): PASAJES_MANAGER / ADMIN_MANAGER / staff.
        - `list` / `retrieve` / `export` (full or near-full RRHH data): RRHH-privileged
          roles only (IsHRManager) — non-RRHH authenticated users get 403.
        - Write actions (create/update/partial_update/destroy): staff only.
        """
        if self.action == 'buscar_dni':
            return [IsAuthenticated(), IsPasajesStaff()]
        if self.action in ('list', 'retrieve', 'export'):
            return [IsAuthenticated(), IsHRManager()]
        if not self.request.user.is_staff:
            from rest_framework.permissions import IsAdminUser
            return [IsAuthenticated(), IsAdminUser()]
        return [IsAuthenticated()]

    # ------------------------------------------------------------------
    # Extra actions
    # ------------------------------------------------------------------

    @extend_schema(
        tags=['personal'],
        summary='Search personnel by exact DNI',
        parameters=[OpenApiParameter('dni', str, OpenApiParameter.QUERY, required=True)],
    )
    @action(detail=False, methods=['get'], url_path='buscar-dni')
    def buscar_dni(self, request):
        """
        Return a single Personal record matching the given DNI exactly.
        Returns 404 if not found.
        Used by the pasajes module to auto-fill traveller data.
        """
        dni = request.query_params.get('dni', '').strip()
        if not dni:
            return Response(
                {'detail': 'El parametro "dni" es requerido.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            persona = Personal.objects.select_related('proyecto', 'user').get(dni=dni)
        except Personal.DoesNotExist:
            return Response(
                {'detail': f'No se encontro personal con DNI {dni}.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        # SYSPCC-006 FIX 1: only non-sensitive fields — this endpoint is reachable
        # by PASAJES_MANAGER, which must never see salary/banking/address data.
        serializer = PersonalDNILookupSerializer(persona, context={'request': request})
        return Response(serializer.data)

    @extend_schema(
        tags=['personal'],
        summary='Export personnel list to Excel',
    )
    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        """
        Generate and return an Excel file with the current filtered queryset.
        Applies the same filters as the list endpoint.
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return Response(
                {'detail': 'openpyxl no esta instalado en el servidor.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        queryset = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Personal'

        # Header row
        headers = [
            'ID', 'DNI', 'Apellidos y Nombres', 'Sexo', 'Estado Civil',
            'Fecha Nacimiento', 'Edad', 'Celular', 'Email Personal',
            'Estado', 'Fecha Ingreso', 'Fecha Cese',
            'Proyecto', 'Sede', 'Puesto', 'Guardia', 'Condicion',
            'N Fotocheck',
        ]
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_idx, persona in enumerate(queryset, start=2):
            ws.cell(row=row_idx, column=1, value=persona.pk)
            ws.cell(row=row_idx, column=2, value=persona.dni)
            ws.cell(row=row_idx, column=3, value=persona.apellidos_nombres)
            ws.cell(row=row_idx, column=4, value=persona.get_sexo_display() if persona.sexo else '')
            ws.cell(row=row_idx, column=5, value=persona.get_estado_civil_display() if persona.estado_civil else '')
            ws.cell(row=row_idx, column=6, value=persona.fecha_nacimiento)
            ws.cell(row=row_idx, column=7, value=persona.edad)
            ws.cell(row=row_idx, column=8, value=persona.celular)
            ws.cell(row=row_idx, column=9, value=persona.email_personal)
            ws.cell(row=row_idx, column=10, value=persona.get_estado_display())
            ws.cell(row=row_idx, column=11, value=persona.fecha_ingreso)
            ws.cell(row=row_idx, column=12, value=persona.fecha_cese)
            ws.cell(row=row_idx, column=13, value=persona.proyecto.name if persona.proyecto else '')
            ws.cell(row=row_idx, column=14, value=persona.sede)
            ws.cell(row=row_idx, column=15, value=persona.puesto)
            ws.cell(row=row_idx, column=16, value=persona.get_guardia_display() if persona.guardia else '')
            ws.cell(row=row_idx, column=17, value=persona.get_condicion_trabajo_display() if persona.condicion_trabajo else '')
            ws.cell(row=row_idx, column=18, value=persona.numero_fotocheck)

        # Auto-width columns
        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="personal.xlsx"'
        return response
