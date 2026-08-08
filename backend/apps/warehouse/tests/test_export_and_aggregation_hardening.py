"""
SYSPCC-013 FIX 1 / FIX 2 / FIX 3 — warehouse module.

FIX 1: Inventory/Movement Excel exports must escape formula-injection
       payloads in free-text fields (description, location, brand,
       supplier/destination) via apps.core.utils.excel.sanitize_excel_value.
FIX 2: exports must cap at MAX_EXPORT_ROWS with a visible truncation warning
       instead of building an unbounded workbook.
FIX 3: `alerts()` / `summary()` must compute total stock with a DB
       aggregation (Sum('stocks__quantity')), the same pattern already used
       by `_apply_inventory_filters`, instead of iterating
       `item.stocks.all()` in Python. These tests assert correctness against
       multiple InventoryStock rows per item — the number that must match
       whether it's computed in Python or in the DB.
"""
import io
from decimal import Decimal
from unittest import mock

import openpyxl
import pytest
from rest_framework import status
from rest_framework.test import APIClient

from apps.core.enums import ItemTypeChoices, WarehouseOriginChoices
from apps.warehouse.models import Inventory, InventoryStock
from apps.warehouse.services import movements as warehouse_svc


def _client_for(user):
    client = APIClient()
    client.force_authenticate(user=user)
    return client


def _read_workbook(response):
    return openpyxl.load_workbook(io.BytesIO(response.content))


@pytest.fixture
def inventory_item(db):
    return Inventory.objects.create(
        product_code='CEM-001',
        description='Cemento Portland Tipo I',
        unit='bls',
        item_type=ItemTypeChoices.MATERIAL,
        min_stock=Decimal('5.000'),
    )


@pytest.mark.django_db
class TestInventoryExportFormulaInjection:
    def test_malicious_description_is_escaped(self, logistics_coordinator):
        Inventory.objects.create(
            product_code='INJ-001',
            description='=cmd(calc)',
            unit='und',
            location='+A1',
        )
        client = _client_for(logistics_coordinator)

        response = client.get('/api/v1/warehouse/inventory/export/')

        assert response.status_code == status.HTTP_200_OK
        wb = _read_workbook(response)
        ws = wb.active
        row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        # headers: Codigo(1) Descripcion(2) Unidad(3) ... Ubicacion(8)
        assert row[1] == "'=cmd(calc)"
        assert row[7] == "'+A1"


@pytest.mark.django_db
class TestInventoryExportTruncation:
    def test_export_truncates_past_the_cap(self, logistics_coordinator):
        for i in range(3):
            Inventory.objects.create(product_code=f'ITM-{i:03d}', description=f'Item {i}', unit='und')
        client = _client_for(logistics_coordinator)

        with mock.patch('apps.core.utils.excel.MAX_EXPORT_ROWS', 1):
            response = client.get('/api/v1/warehouse/inventory/export/')

        assert response.status_code == status.HTTP_200_OK
        wb = _read_workbook(response)
        ws = wb.active
        assert 'ADVERTENCIA' in ws.cell(row=1, column=1).value
        assert ws.cell(row=2, column=1).value == 'Codigo'


@pytest.mark.django_db
class TestMovementExportFormulaInjection:
    def test_malicious_supplier_name_is_escaped(self, logistics_coordinator, inventory_item, user):
        warehouse_svc.register_entry(
            inventory=inventory_item,
            quantity=Decimal('10'),
            warehouse=WarehouseOriginChoices.CENTRAL,
            project_id=None,
            source_type='PURCHASE',
            supplier_name='=cmd()|calc',
            invoice_number='F001-1',
            notes='',
            registered_by=user,
        )
        client = _client_for(logistics_coordinator)

        response = client.get('/api/v1/warehouse/movements/export/')

        assert response.status_code == status.HTTP_200_OK
        wb = _read_workbook(response)
        ws = wb.active
        row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        # headers: N.Movimiento(1) Tipo(2) Fecha(3) Codigo(4) Descripcion(5)
        # Cantidad(6) Unidad(7) Almacen(8) Proveedor/Destino(9)
        assert row[8] == "'=cmd()|calc"


@pytest.mark.django_db
class TestInventoryAlertsAggregation:
    def test_alerts_sums_stock_across_multiple_warehouses(self, logistics_coordinator, inventory_item):
        """Two InventoryStock rows for the same item (CENTRAL + SITE) must be
        summed together — total 3 <= min_stock 5 => it's a low-stock alert."""
        InventoryStock.objects.create(
            inventory=inventory_item, warehouse_type=WarehouseOriginChoices.CENTRAL,
            project=None, department=None, quantity=Decimal('2.000'),
        )
        InventoryStock.objects.create(
            inventory=inventory_item, warehouse_type=WarehouseOriginChoices.SITE,
            project=None, department=None, quantity=Decimal('1.000'),
        )
        client = _client_for(logistics_coordinator)

        response = client.get('/api/v1/warehouse/inventory/alerts/')

        assert response.status_code == status.HTTP_200_OK
        alert = next(a for a in response.data if a['id'] == inventory_item.pk)
        assert alert['current_stock'] == 3.0
        assert alert['deficit'] == 2.0

    def test_alerts_excludes_items_above_threshold(self, logistics_coordinator, inventory_item):
        InventoryStock.objects.create(
            inventory=inventory_item, warehouse_type=WarehouseOriginChoices.CENTRAL,
            project=None, department=None, quantity=Decimal('100.000'),
        )
        client = _client_for(logistics_coordinator)

        response = client.get('/api/v1/warehouse/inventory/alerts/')

        assert response.status_code == status.HTTP_200_OK
        assert all(a['id'] != inventory_item.pk for a in response.data)

    def test_alerts_includes_item_with_no_stock_rows_at_all(self, logistics_coordinator, inventory_item):
        """An item with zero InventoryStock rows must still be flagged (0 <=
        min_stock) — the Coalesce(Sum(...), 0) must behave like the old
        `sum(... for empty iterable) == 0`, not drop the item via an inner
        join."""
        client = _client_for(logistics_coordinator)

        response = client.get('/api/v1/warehouse/inventory/alerts/')

        assert response.status_code == status.HTTP_200_OK
        alert = next(a for a in response.data if a['id'] == inventory_item.pk)
        assert alert['current_stock'] == 0.0


@pytest.mark.django_db
class TestInventorySummaryAggregation:
    def test_summary_low_stock_count_matches_multi_warehouse_totals(self, logistics_coordinator, inventory_item):
        InventoryStock.objects.create(
            inventory=inventory_item, warehouse_type=WarehouseOriginChoices.CENTRAL,
            project=None, department=None, quantity=Decimal('2.000'),
        )
        InventoryStock.objects.create(
            inventory=inventory_item, warehouse_type=WarehouseOriginChoices.SITE,
            project=None, department=None, quantity=Decimal('1.000'),
        )
        other_item = Inventory.objects.create(
            product_code='OK-001', description='Bien abastecido', unit='und',
            min_stock=Decimal('1.000'),
        )
        InventoryStock.objects.create(
            inventory=other_item, warehouse_type=WarehouseOriginChoices.CENTRAL,
            project=None, department=None, quantity=Decimal('50.000'),
        )
        client = _client_for(logistics_coordinator)

        response = client.get('/api/v1/warehouse/inventory/summary/')

        assert response.status_code == status.HTTP_200_OK
        assert response.data['total_items'] == 2
        assert response.data['low_stock_count'] == 1
