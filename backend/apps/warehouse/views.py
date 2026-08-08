"""
Warehouse API views — standalone inventory management module.
Completely independent from the RQ system.

Only users with CENTRAL_WAREHOUSE or SITE_WAREHOUSE roles may access these endpoints.
"""
from django.db.models import Q, Sum, DecimalField, Value, F
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated, BasePermission
from rest_framework.response import Response

from apps.warehouse.models import Inventory, InventoryStock, InventoryMovement, MovementGroup
from apps.warehouse.serializers import (
    InventoryListSerializer,
    InventoryCreateUpdateSerializer,
    MovementListSerializer,
    MovementGroupSerializer,
    EntryCreateSerializer,
    ExitCreateSerializer,
    BatchEntryCreateSerializer,
    BatchExitCreateSerializer,
)
from apps.warehouse.services import movements as warehouse_svc
from apps.warehouse.services.onedrive import OneDriveService
from apps.warehouse.services.pdf_generator import generate_movement_pdf, generate_group_pdf


# ---------------------------------------------------------------------------
# Pagination
# ---------------------------------------------------------------------------

class WarehousePagination(PageNumberPagination):
    page_size = 25
    page_size_query_param = 'page_size'
    max_page_size = 100


# ---------------------------------------------------------------------------
# Permission
# ---------------------------------------------------------------------------

class IsWarehouseRole(BasePermission):
    """Allow access to warehouse and logistics users."""

    message = 'Se requiere el rol de Almacén, Logística o similar.'

    ALLOWED_ROLES = [
        'CENTRAL_WAREHOUSE', 'SITE_WAREHOUSE',
        'LOGISTICS_COORDINATOR', 'LOGISTICS_SUPERVISOR', 'LOGISTICS_CHIEF',
    ]

    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        return request.user.user_roles.filter(role__in=self.ALLOWED_ROLES).exists()


# ---------------------------------------------------------------------------
# Inventory ViewSet
# ---------------------------------------------------------------------------

class InventoryViewSet(viewsets.ModelViewSet):
    """Full CRUD for the inventory catalog plus auxiliary read actions."""

    permission_classes = [IsAuthenticated, IsWarehouseRole]
    pagination_class = WarehousePagination
    queryset = Inventory.objects.prefetch_related('stocks', 'stocks__project').all()

    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return InventoryCreateUpdateSerializer
        return InventoryListSerializer

    # -- shared filter helper ------------------------------------------------

    def _apply_inventory_filters(self, qs, params):
        """Apply common query filters to an inventory queryset."""
        category = params.get('category')
        item_type = params.get('item_type')
        search = params.get('search')
        low_stock = params.get('low_stock')

        if category:
            qs = qs.filter(category__iexact=category)
        if item_type:
            qs = qs.filter(item_type=item_type)
        if search:
            qs = qs.filter(
                Q(product_code__icontains=search) |
                Q(description__icontains=search) |
                Q(brand__icontains=search)
            )

        # Annotate total stock at the database level for filtering
        qs = qs.annotate(
            total_stock=Coalesce(
                Sum('stocks__quantity'),
                Value(0),
                output_field=DecimalField(),
            )
        )
        if low_stock == 'true':
            qs = qs.filter(total_stock__lte=F('min_stock'))

        return qs

    # -- list with filters ---------------------------------------------------

    def list(self, request):
        qs = self._apply_inventory_filters(self.get_queryset(), request.query_params)

        # Support no_page=true for dropdowns that need the full list
        if request.query_params.get('no_page') == 'true':
            serializer = InventoryListSerializer(qs, many=True)
            return Response(serializer.data)

        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = InventoryListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = InventoryListSerializer(qs, many=True)
        return Response(serializer.data)

    # -- kardex (movement history for one item) ------------------------------

    @action(detail=True, methods=['get'])
    def kardex(self, request, pk=None):
        """Return all movements (kardex) for a specific inventory item."""
        item = self.get_object()
        movements = InventoryMovement.objects.filter(
            inventory=item
        ).select_related('registered_by', 'requested_by', 'authorized_by', 'project')

        warehouse = request.query_params.get('warehouse')
        if warehouse:
            movements = movements.filter(warehouse=warehouse)

        return Response({
            'item': InventoryListSerializer(item).data,
            'movements': MovementListSerializer(movements, many=True).data,
        })

    # -- recent movements for one item --------------------------------------

    @action(detail=True, methods=['get'], url_path='recent-movements')
    def recent_movements(self, request, pk=None):
        """Return the 5 most recent movements for a specific inventory item."""
        item = self.get_object()
        movements = InventoryMovement.objects.filter(inventory=item).select_related(
            'registered_by'
        ).order_by('-created_at')[:5]
        return Response(MovementListSerializer(movements, many=True).data)

    # -- categories dropdown -------------------------------------------------

    @action(detail=False, methods=['get'])
    def categories(self, request):
        """Return distinct non-empty categories for filter dropdowns."""
        cats = (
            Inventory.objects
            .exclude(category='')
            .values_list('category', flat=True)
            .distinct()
            .order_by('category')
        )
        return Response(list(cats))

    # -- stock alerts --------------------------------------------------------

    @action(detail=False, methods=['get'])
    def alerts(self, request):
        """Return items whose total stock is at or below their minimum threshold."""
        items = Inventory.objects.prefetch_related('stocks').all()
        result = []
        for item in items:
            total = sum(s.quantity for s in item.stocks.all())
            if total <= item.min_stock:
                result.append({
                    'id': item.id,
                    'product_code': item.product_code,
                    'description': item.description,
                    'unit': item.unit,
                    'min_stock': float(item.min_stock),
                    'current_stock': float(total),
                    'deficit': float(item.min_stock - total),
                })
        return Response(result)

    # -- stock check (used by RQ logistics) ----------------------------------

    @action(detail=False, methods=['get'], url_path='check-stock')
    def check_stock(self, request):
        """
        Check availability for a comma-separated list of item descriptions.
        Used by the RQ logistics module; kept read-only and open to IsAuthenticated only.
        Query params: project_id, descriptions (comma-separated)
        """
        project_id = request.query_params.get('project_id')
        descriptions = request.query_params.get('descriptions', '')

        if not descriptions:
            return Response([])

        desc_list = [d.strip() for d in descriptions.split(',') if d.strip()]
        results = []

        for desc in desc_list:
            inventory = Inventory.objects.filter(
                Q(description__icontains=desc) | Q(product_code__icontains=desc)
            ).first()

            if inventory:
                central_stock = InventoryStock.objects.filter(
                    inventory=inventory, warehouse_type='CENTRAL'
                ).first()
                site_stock = InventoryStock.objects.filter(
                    inventory=inventory, warehouse_type='SITE', project_id=project_id
                ).first() if project_id else None

                results.append({
                    'description': desc,
                    'matched_product': inventory.product_code,
                    'matched_description': inventory.description,
                    'unit': inventory.unit,
                    'stock_central': float(central_stock.quantity) if central_stock else 0,
                    'stock_obra': float(site_stock.quantity) if site_stock else 0,
                    'min_stock': float(inventory.min_stock),
                    'has_stock': (
                        (central_stock and central_stock.quantity > 0)
                        or (site_stock and site_stock.quantity > 0)
                    ),
                })
            else:
                results.append({
                    'description': desc,
                    'matched_product': None,
                    'matched_description': None,
                    'unit': None,
                    'stock_central': 0,
                    'stock_obra': 0,
                    'min_stock': 0,
                    'has_stock': False,
                })

        return Response(results)

    # -- dashboard summary ---------------------------------------------------

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Quick stats for the warehouse dashboard."""
        today = timezone.now().date()

        total_items = Inventory.objects.count()
        items = Inventory.objects.prefetch_related('stocks').all()
        low_stock_count = sum(
            1 for item in items
            if sum(s.quantity for s in item.stocks.all()) <= item.min_stock
        )
        today_entries = InventoryMovement.objects.filter(
            movement_type='ENTRY', created_at__date=today
        ).count()
        today_exits = InventoryMovement.objects.filter(
            movement_type='EXIT', created_at__date=today
        ).count()

        return Response({
            'total_items': total_items,
            'low_stock_count': low_stock_count,
            'today_entries': today_entries,
            'today_exits': today_exits,
        })

    # -- Excel export --------------------------------------------------------

    @action(detail=False, methods=['get'], url_path='export')
    def export_excel(self, request):
        """Export the current filtered inventory list as an xlsx file."""
        qs = self._apply_inventory_filters(
            Inventory.objects.prefetch_related('stocks').all(),
            request.query_params,
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Inventario'

        headers = [
            'Codigo', 'Descripcion', 'Unidad', 'Categoria', 'Tipo',
            'Marca', 'Modelo', 'Ubicacion', 'Stock Total', 'Stock Minimo',
        ]
        ws.append(headers)

        # Bold header row
        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold

        for item in qs:
            total = float(sum(s.quantity for s in item.stocks.all()))
            ws.append([
                item.product_code,
                item.description,
                item.unit,
                item.category,
                item.get_item_type_display(),
                item.brand,
                item.model_name,
                item.location,
                total,
                float(item.min_stock),
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="inventario.xlsx"'
        wb.save(response)
        return response


# ---------------------------------------------------------------------------
# Movement ViewSet
# ---------------------------------------------------------------------------

class MovementViewSet(viewsets.ReadOnlyModelViewSet):
    """
    List/retrieve movements plus POST actions for entry and exit.
    Write actions (entry, exit) require warehouse role.
    """

    permission_classes = [IsAuthenticated, IsWarehouseRole]
    serializer_class = MovementListSerializer
    pagination_class = WarehousePagination
    queryset = InventoryMovement.objects.select_related(
        'inventory', 'registered_by', 'requested_by', 'authorized_by', 'project'
    ).all()

    # -- shared filter helper ------------------------------------------------

    def _apply_movement_filters(self, qs, params):
        """Apply common query filters to a movement queryset."""
        movement_type = params.get('movement_type')
        warehouse = params.get('warehouse')
        search = params.get('search')
        date_from = params.get('date_from')
        date_to = params.get('date_to')
        supplier = params.get('supplier')
        project = params.get('project')
        qty_min = params.get('qty_min')
        qty_max = params.get('qty_max')

        if movement_type:
            qs = qs.filter(movement_type=movement_type)
        if warehouse:
            qs = qs.filter(warehouse=warehouse)
        if search:
            qs = qs.filter(
                Q(movement_number__icontains=search) |
                Q(inventory__product_code__icontains=search) |
                Q(inventory__description__icontains=search)
            )
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        if supplier:
            qs = qs.filter(supplier_name__icontains=supplier)
        if project:
            qs = qs.filter(project_id=project)
        if qty_min:
            qs = qs.filter(quantity__gte=qty_min)
        if qty_max:
            qs = qs.filter(quantity__lte=qty_max)

        return qs

    def list(self, request):
        qs = self._apply_movement_filters(self.get_queryset(), request.query_params)

        # Support no_page=true for full list without pagination
        if request.query_params.get('no_page') == 'true':
            serializer = self.get_serializer(qs, many=True)
            return Response(serializer.data)

        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    # -- entry ---------------------------------------------------------------

    @action(detail=False, methods=['post'])
    def entry(self, request):
        """Register a new inventory entry (entrada)."""
        serializer = EntryCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        d = serializer.validated_data
        try:
            movement = warehouse_svc.register_entry(
                inventory=d['inventory'],
                quantity=d['quantity'],
                warehouse=d['warehouse'],
                project_id=d.get('project'),
                source_type=d['source_type'],
                supplier_name=d.get('supplier_name', ''),
                invoice_number=d.get('invoice_number', ''),
                notes=d.get('notes', ''),
                registered_by=request.user,
            )
            return Response(
                MovementListSerializer(movement).data,
                status=status.HTTP_201_CREATED,
            )
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    # -- exit ----------------------------------------------------------------

    @action(detail=False, methods=['post'])
    def exit(self, request):
        """Register a new inventory exit (salida)."""
        serializer = ExitCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        d = serializer.validated_data
        try:
            movement = warehouse_svc.register_exit(
                inventory=d['inventory'],
                quantity=d['quantity'],
                warehouse=d['warehouse'],
                project_id=d.get('project'),
                destination_type=d['destination_type'],
                destination_detail=d.get('destination_detail', ''),
                requested_by_id=d.get('requested_by'),
                authorized_by_id=d.get('authorized_by'),
                notes=d.get('notes', ''),
                registered_by=request.user,
            )
            return Response(
                MovementListSerializer(movement).data,
                status=status.HTTP_201_CREATED,
            )
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    # -- batch entry ---------------------------------------------------------

    @action(detail=False, methods=['post'], url_path='batch-entry')
    def batch_entry(self, request):
        """Register multiple inventory entries as a single vale (batch entrada)."""
        serializer = BatchEntryCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        d = serializer.validated_data

        items_data = [
            {
                'inventory_id': item['inventory'].pk,
                'quantity': item['quantity'],
                'line': idx,
            }
            for idx, item in enumerate(d['items'], start=1)
        ]

        try:
            group = warehouse_svc.register_entry_batch(
                items_data=items_data,
                warehouse=d['warehouse'],
                project_id=d.get('project'),
                source_type=d['source_type'],
                supplier_name=d.get('supplier_name', ''),
                invoice_number=d.get('invoice_number', ''),
                notes=d.get('notes', ''),
                registered_by=request.user,
            )
            # Generate PDF and upload to OneDrive (sync — user waits)
            from apps.warehouse.services.movements import _upload_group_pdf_sync
            _upload_group_pdf_sync(group.pk)
            group.refresh_from_db()

            return Response(
                MovementGroupSerializer(group).data,
                status=status.HTTP_201_CREATED,
            )
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    # -- batch exit ----------------------------------------------------------

    @action(detail=False, methods=['post'], url_path='batch-exit')
    def batch_exit(self, request):
        """Register multiple inventory exits as a single vale (batch salida)."""
        serializer = BatchExitCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        d = serializer.validated_data

        items_data = [
            {
                'inventory_id': item['inventory'].pk,
                'quantity': item['quantity'],
                'line': idx,
            }
            for idx, item in enumerate(d['items'], start=1)
        ]

        try:
            group = warehouse_svc.register_exit_batch(
                items_data=items_data,
                warehouse=d['warehouse'],
                project_id=d.get('project'),
                destination_type=d['destination_type'],
                destination_detail=d.get('destination_detail', ''),
                requested_by_id=d.get('requested_by'),
                authorized_by_id=d.get('authorized_by'),
                notes=d.get('notes', ''),
                registered_by=request.user,
            )
            # Generate PDF and upload to OneDrive (sync — user waits)
            from apps.warehouse.services.movements import _upload_group_pdf_sync
            _upload_group_pdf_sync(group.pk)
            group.refresh_from_db()

            return Response(
                MovementGroupSerializer(group).data,
                status=status.HTTP_201_CREATED,
            )
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    # -- voucher (vale de salida) --------------------------------------------

    @action(detail=True, methods=['get'])
    def voucher(self, request, pk=None):
        """
        Return movement data formatted for printing a voucher.
        If the movement belongs to a group, the response includes all group items
        so the frontend can render a multi-item vale.
        """
        movement = self.get_object()

        # Shared fields
        data = {
            'movement_number': movement.movement_number,
            'movement_type': movement.movement_type,
            'movement_type_display': movement.get_movement_type_display(),
            'date': movement.created_at.strftime('%d/%m/%Y %H:%M'),
            'warehouse': movement.get_warehouse_display(),
            'project': movement.project.name if movement.project else '',
            'project_code': movement.project.code if movement.project else '',
            'destination_type': movement.destination_type,
            'destination_detail': movement.destination_detail,
            'supplier_name': movement.supplier_name,
            'invoice_number': movement.invoice_number,
            'source_type': movement.source_type,
            'requested_by': (
                movement.requested_by.get_full_name() if movement.requested_by else ''
            ),
            'authorized_by': (
                movement.authorized_by.get_full_name() if movement.authorized_by else ''
            ),
            'registered_by': movement.registered_by.get_full_name(),
            'notes': movement.notes,
            'document_url': movement.document_url or '',
        }

        # If this movement is part of a group, surface all items
        if movement.group_id:
            group = (
                MovementGroup.objects
                .prefetch_related('movements__inventory')
                .get(pk=movement.group_id)
            )
            data['group_number'] = group.group_number
            data['document_url'] = group.document_url or movement.document_url or ''
            data['items'] = [
                {
                    'product_code': mv.inventory.product_code,
                    'description': mv.inventory.description,
                    'unit': mv.inventory.unit,
                    'quantity': float(mv.quantity),
                }
                for mv in group.movements.select_related('inventory').order_by('movement_number')
            ]
        else:
            data['group_number'] = None
            data['items'] = [
                {
                    'product_code': movement.inventory.product_code,
                    'description': movement.inventory.description,
                    'unit': movement.inventory.unit,
                    'quantity': float(movement.quantity),
                }
            ]

        return Response(data)

    # -- group voucher -------------------------------------------------------

    @action(
        detail=False,
        methods=['get'],
        url_path=r'groups/(?P<group_pk>[0-9]+)/voucher',
    )
    def group_voucher(self, request, group_pk=None):
        """Return full voucher data for a MovementGroup (multi-item)."""
        try:
            group = (
                MovementGroup.objects
                .select_related('registered_by', 'project')
                .prefetch_related('movements__inventory')
                .get(pk=group_pk)
            )
        except MovementGroup.DoesNotExist:
            return Response({'detail': 'Grupo no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        return Response(MovementGroupSerializer(group).data)

    # -- group PDF download --------------------------------------------------

    @action(
        detail=False,
        methods=['get'],
        url_path=r'groups/(?P<group_pk>[0-9]+)/pdf',
    )
    def group_pdf(self, request, group_pk=None):
        """Generate and stream the multi-item PDF for a MovementGroup."""
        try:
            group = (
                MovementGroup.objects
                .select_related('registered_by', 'project')
                .prefetch_related('movements__inventory')
                .get(pk=group_pk)
            )
        except MovementGroup.DoesNotExist:
            return Response({'detail': 'Grupo no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        from apps.warehouse.services.pdf_generator import get_group_filename
        pdf_bytes = generate_group_pdf(group)
        filename = get_group_filename(group)
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response

    # -- download PDF -------------------------------------------------------

    @action(detail=True, methods=['get'], url_path='pdf')
    def download_pdf(self, request, pk=None):
        """Generate and return the movement PDF for download."""
        movement = self.get_object()
        pdf_bytes = generate_movement_pdf(movement)

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        from apps.warehouse.services.pdf_generator import get_movement_filename
        response['Content-Disposition'] = f'inline; filename="{get_movement_filename(movement)}"'
        return response

    # -- Excel export --------------------------------------------------------

    @action(detail=False, methods=['get'], url_path='export')
    def export_excel(self, request):
        """Export the current filtered movement list as an xlsx file."""
        qs = self._apply_movement_filters(
            InventoryMovement.objects.select_related(
                'inventory', 'registered_by', 'project'
            ).all(),
            request.query_params,
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Movimientos'

        headers = [
            'N. Movimiento', 'Tipo', 'Fecha', 'Codigo', 'Descripcion',
            'Cantidad', 'Unidad', 'Almacen', 'Proveedor/Destino', 'Registrado por',
        ]
        ws.append(headers)

        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold

        for mv in qs:
            provider_dest = mv.supplier_name or mv.destination_detail or ''
            ws.append([
                mv.movement_number,
                mv.get_movement_type_display(),
                mv.created_at.strftime('%d/%m/%Y %H:%M'),
                mv.inventory.product_code,
                mv.inventory.description,
                float(mv.quantity),
                mv.inventory.unit,
                mv.get_warehouse_display(),
                provider_dest,
                mv.registered_by.get_full_name() if mv.registered_by else '',
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="movimientos.xlsx"'
        wb.save(response)
        return response


# ---------------------------------------------------------------------------
# OneDrive Integration Views
# ---------------------------------------------------------------------------

class IsAdminUser(BasePermission):
    """Only superusers / staff can manage OneDrive connection."""
    message = 'Solo el administrador puede configurar OneDrive.'

    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        # Status is readable by anyone authenticated, write actions need admin
        if view.action == 'status':
            return True
        return request.user.is_superuser or request.user.is_staff


class OneDriveViewSet(viewsets.ViewSet):
    """
    Endpoints for OneDrive OAuth2 device code flow.
    Only admin/superuser can connect/disconnect. Status is readable by all.
    """

    permission_classes = [IsAuthenticated, IsAdminUser]

    @action(detail=False, methods=['get'])
    def status(self, request):
        """Check if OneDrive is connected."""
        return Response(OneDriveService.get_connection_info())

    @action(detail=False, methods=['post'], url_path='connect')
    def connect(self, request):
        """
        Start the device code flow.
        Returns the user_code and verification_uri for the user to complete auth.
        """
        try:
            data = OneDriveService.initiate_device_code()
            return Response({
                'user_code': data.get('user_code'),
                'verification_uri': data.get('verification_uri'),
                'device_code': data.get('device_code'),
                'expires_in': data.get('expires_in', 900),
                'interval': data.get('interval', 5),
                'message': data.get('message', ''),
            })
        except Exception as exc:
            return Response(
                {'detail': f'Error al iniciar conexion: {str(exc)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=['post'], url_path='poll')
    def poll(self, request):
        """
        Poll for token after user enters the device code.
        Called from the frontend at intervals until auth completes.
        """
        device_code = request.data.get('device_code')
        if not device_code:
            return Response(
                {'detail': 'device_code requerido'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Single poll attempt (don't block the request)
            import requests as http_requests
            resp = http_requests.post(
                'https://login.microsoftonline.com/consumers/oauth2/v2.0/token',
                data={
                    'client_id': '14d82eec-204b-4c2f-b7e8-296a70dab67e',
                    'grant_type': 'urn:ietf:params:oauth:grant-type:device_code',
                    'device_code': device_code,
                },
                timeout=10,
            )
            data = resp.json()

            if 'access_token' in data:
                from datetime import timedelta
                # Get account info
                account_name = ''
                try:
                    me = http_requests.get(
                        'https://graph.microsoft.com/v1.0/me',
                        headers={'Authorization': f'Bearer {data["access_token"]}'},
                        timeout=10,
                    ).json()
                    account_name = me.get('userPrincipalName') or me.get('displayName', '')
                except Exception:
                    pass

                from apps.warehouse.models import OneDriveToken
                expires_at = timezone.now() + timedelta(seconds=data.get('expires_in', 3600))
                OneDriveToken.save_token(
                    access_token=data['access_token'],
                    refresh_token=data.get('refresh_token', ''),
                    expires_at=expires_at,
                    account_name=account_name,
                )
                return Response({
                    'status': 'connected',
                    'account_name': account_name,
                })

            error = data.get('error', '')
            if error == 'authorization_pending':
                return Response({'status': 'pending'})
            if error == 'slow_down':
                return Response({'status': 'pending', 'slow_down': True})

            return Response(
                {'status': 'failed', 'error': error},
                status=status.HTTP_400_BAD_REQUEST,
            )

        except Exception as exc:
            return Response(
                {'detail': str(exc)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=['post'], url_path='disconnect')
    def disconnect(self, request):
        """Remove stored OneDrive tokens."""
        OneDriveService.disconnect()
        return Response({'status': 'disconnected'})


# ---------------------------------------------------------------------------
# Pending Requests for Warehouse — RQ integration
# ---------------------------------------------------------------------------

# States where CENTRAL_WAREHOUSE must act
CENTRAL_PENDING_STATUSES = [
    'IN_STOCK',            # Must dispatch from central stock
    'DISPATCHED_TO_SITE',  # ADM: must update warehouse records
    'DELIVERED',           # ADM: must mark warehouse records updated
]

# States where SITE_WAREHOUSE must act
SITE_PENDING_STATUSES = [
    'DISPATCHED_TO_SITE',  # OPS: must confirm receipt at site
]

STATUS_LABELS = {
    'IN_STOCK':           'Stock disponible — pendiente despacho',
    'DISPATCHED_TO_SITE': 'Despachado — pendiente actualizar registros',
    'DELIVERED':          'Entregado — pendiente confirmar en almacén',
}

PENDING_ACTIONS = {
    'IN_STOCK':           'DISPATCHED',
    'DISPATCHED_TO_SITE': 'WAREHOUSE_RECORDS_UPDATED',
    'DELIVERED':          'WAREHOUSE_RECORDS_UPDATED',
}


class PendingRequestsViewSet(viewsets.ViewSet):
    """
    RQs that require action from the warehouse user.

    GET /warehouse/pending-requests/list/   — paginated list
    GET /warehouse/pending-requests/count/  — badge count by status
    """

    permission_classes = [IsAuthenticated, IsWarehouseRole]

    def _get_pending_statuses(self, user):
        roles = list(user.user_roles.values_list('role', flat=True))
        statuses = set()
        if 'CENTRAL_WAREHOUSE' in roles:
            statuses.update(CENTRAL_PENDING_STATUSES)
        if 'SITE_WAREHOUSE' in roles:
            statuses.update(SITE_PENDING_STATUSES)
        if any(r in roles for r in ['LOGISTICS_COORDINATOR', 'LOGISTICS_SUPERVISOR', 'LOGISTICS_CHIEF']):
            statuses.update(CENTRAL_PENDING_STATUSES)
            statuses.update(SITE_PENDING_STATUSES)
        return list(statuses)

    def _build_queryset(self, user, params):
        from apps.rq.models import Request
        pending_statuses = self._get_pending_statuses(user)
        qs = (
            Request.objects
            .filter(status__in=pending_statuses)
            .select_related('project', 'department', 'requested_by')
            .prefetch_related('items')
            .order_by('fecha_necesidad', '-created_at')
        )
        if params.get('flow'):
            qs = qs.filter(flow=params['flow'])
        if params.get('status'):
            qs = qs.filter(status=params['status'])
        if params.get('search'):
            s = params['search']
            qs = qs.filter(
                Q(rq_number__icontains=s) |
                Q(description__icontains=s) |
                Q(project__name__icontains=s) |
                Q(department__name__icontains=s) |
                Q(requested_by__first_name__icontains=s) |
                Q(requested_by__last_name__icontains=s)
            )
        return qs

    @action(detail=False, methods=['get'], url_path='list')
    def list_pending(self, request):
        """Paginated list of RQs pending warehouse action."""
        qs = self._build_queryset(request.user, request.query_params)
        paginator = WarehousePagination()
        page = paginator.paginate_queryset(qs, request)
        items = page if page is not None else qs

        data = []
        for rq in items:
            data.append({
                'id': rq.id,
                'rq_number': rq.rq_number,
                'flow': rq.flow,
                'status': rq.status,
                'status_label': STATUS_LABELS.get(rq.status, rq.status),
                'pending_action': PENDING_ACTIONS.get(rq.status, ''),
                'priority': rq.priority,
                'fecha_necesidad': rq.fecha_necesidad,
                'fecha_estimada_entrega': rq.fecha_estimada_entrega,
                'project': rq.project.name if rq.project else None,
                'project_code': rq.project.code if rq.project else None,
                'department': rq.department.name if rq.department else None,
                'requested_by': rq.requested_by.get_full_name(),
                'items_count': rq.items.count(),
                'stock_items': rq.items.filter(supply_source='STOCK').count(),
                'purchase_items': rq.items.filter(supply_source='PURCHASE').count(),
                'acquisition_type': rq.acquisition_type,
                'created_at': rq.created_at,
            })

        if page is not None:
            return paginator.get_paginated_response(data)
        return Response(data)

    @action(detail=False, methods=['get'], url_path='count')
    def count_pending(self, request):
        """Badge count of pending requests, broken down by status."""
        from django.db.models import Count
        qs = self._build_queryset(request.user, request.query_params)
        breakdown = qs.values('status').annotate(total=Count('id')).order_by('status')
        return Response({
            'total': qs.count(),
            'by_status': [
                {
                    'status': item['status'],
                    'label': STATUS_LABELS.get(item['status'], item['status']),
                    'count': item['total'],
                }
                for item in breakdown
            ],
        })
