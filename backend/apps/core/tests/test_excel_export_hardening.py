"""
SYSPCC-013 FIX 1 / FIX 2: Excel export hardening.

FIX 1 — CSV/Excel formula injection: free-text fields (nombres, detalle,
cargo, razon_social, description, ...) written verbatim into an exported
.xlsx cell let an attacker whose input reaches an export (e.g. a `nombres`
or `detalle` value of `=cmd(...)`) execute a live formula the moment the
file is opened in Excel/Sheets. `apps.core.utils.excel.sanitize_excel_value`
neutralizes this by prefixing a leading apostrophe on any string starting
with `= + - @`.

FIX 2 — unbounded export queryset: building the whole workbook in memory
over an unfiltered queryset risks OOM on a large table.
`apps.core.utils.excel.truncate_for_export` caps rows at MAX_EXPORT_ROWS and
reports whether truncation happened; the chosen behavior across every export
view is "truncate with a visible warning row", not a 400 — these are
filtered report downloads for office/warehouse staff, not paginated data
APIs, so silently refusing the whole file over a hard cap is worse UX than
handing back the first N rows plus a clear notice to narrow the filters.

This file covers the shared helper directly (unit tests) plus one
integration test per export endpoint reachable in this ticket's scope
(personal export, pasajes export) to prove the helper is actually wired in,
not just defined.
"""
import io
from unittest import mock

import openpyxl
import pytest
from rest_framework import status
from rest_framework.test import APIClient

from apps.administracion.models import Pasaje
from apps.core.models import Personal
from apps.core.utils.excel import MAX_EXPORT_ROWS, sanitize_excel_value, truncate_for_export


def _client_for(user):
    client = APIClient()
    client.force_authenticate(user=user)
    return client


def _read_workbook(response):
    return openpyxl.load_workbook(io.BytesIO(response.content))


class TestSanitizeExcelValue:
    @pytest.mark.parametrize('payload', ['=cmd()', '+1+1', '-2+3', '@SUM(A1:A9)'])
    def test_formula_prefixes_get_escaped(self, payload):
        assert sanitize_excel_value(payload) == "'" + payload

    def test_plain_text_is_untouched(self):
        assert sanitize_excel_value('Juan Perez') == 'Juan Perez'

    def test_non_string_values_are_untouched(self):
        assert sanitize_excel_value(123) == 123
        assert sanitize_excel_value(12.5) == 12.5
        assert sanitize_excel_value(None) is None

    def test_empty_string_is_untouched(self):
        assert sanitize_excel_value('') == ''


@pytest.mark.django_db
class TestTruncateForExport:
    def test_queryset_under_limit_is_not_truncated(self):
        Personal.objects.create(dni='11111111', apellidos_nombres='Uno')
        Personal.objects.create(dni='22222222', apellidos_nombres='Dos')

        qs, was_truncated = truncate_for_export(Personal.objects.all())

        assert was_truncated is False
        assert list(qs) == list(Personal.objects.all())

    def test_queryset_over_limit_is_truncated(self):
        for i in range(5):
            Personal.objects.create(dni=f'{i:08d}', apellidos_nombres=f'Persona {i}')

        qs, was_truncated = truncate_for_export(Personal.objects.all(), max_rows=3)

        assert was_truncated is True
        assert len(qs) == 3

    def test_default_max_rows_matches_module_constant(self):
        """`max_rows` defaults to MAX_EXPORT_ROWS looked up at call time, not
        bound into the function signature — confirms the two stay in sync."""
        Personal.objects.create(dni='11111111', apellidos_nombres='Uno')
        with mock.patch('apps.core.utils.excel.MAX_EXPORT_ROWS', 0):
            qs, was_truncated = truncate_for_export(Personal.objects.all())
        assert was_truncated is True
        assert len(qs) == 0
        # Unpatched, the real constant is the documented 10k cap.
        assert MAX_EXPORT_ROWS == 10000


@pytest.mark.django_db
class TestPersonalExportFormulaInjection:
    def test_malicious_field_value_is_escaped_in_exported_workbook(self, admin_manager):
        Personal.objects.create(
            dni='87654321',
            apellidos_nombres='=cmd()',
            celular='+51999999999',
        )
        client = _client_for(admin_manager)

        response = client.get('/api/v1/personal/export/')

        assert response.status_code == status.HTTP_200_OK
        wb = _read_workbook(response)
        ws = wb.active
        # apellidos_nombres is column 3
        row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        assert row[2] == "'=cmd()"
        # celular ('+51999999999') is column 8
        assert row[7] == "'+51999999999"


@pytest.mark.django_db
class TestPersonalExportTruncation:
    def test_export_truncates_and_warns_past_the_cap(self, admin_manager):
        for i in range(3):
            Personal.objects.create(dni=f'{i:08d}', apellidos_nombres=f'Persona {i}')
        client = _client_for(admin_manager)

        with mock.patch('apps.core.utils.excel.MAX_EXPORT_ROWS', 2):
            response = client.get('/api/v1/personal/export/')

        assert response.status_code == status.HTTP_200_OK
        wb = _read_workbook(response)
        ws = wb.active
        # Row 1: warning, row 2: header, rows 3-4: the 2 truncated data rows.
        warning_cell = ws.cell(row=1, column=1).value
        assert 'ADVERTENCIA' in warning_cell
        assert ws.cell(row=2, column=1).value == 'ID'
        data_rows = list(ws.iter_rows(min_row=3, values_only=True))
        assert len(data_rows) == 2


@pytest.mark.django_db
class TestPasajeExportFormulaInjection:
    def test_malicious_detalle_is_escaped_in_exported_workbook(self, admin_manager):
        Pasaje.objects.create(
            tipo='B',
            dni='12345678',
            nombres='=cmd()',
            tipo_trabajador='WORKER',
            detalle='@SUM(1+1)*cmd()',
            creado_por=admin_manager,
        )
        client = _client_for(admin_manager)

        response = client.get('/api/v1/administracion/pasajes/export/')

        assert response.status_code == status.HTTP_200_OK
        wb = _read_workbook(response)
        ws = wb.active
        row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        # headers: ID(1) Tipo(2) FechaBajada(3) EmbarqueBajada(4) DestinoBajada(5)
        # FechaSubida(6) EmbarqueSubida(7) DestinoSubida(8) DNI(9) Nombres(10) ...
        assert row[9] == "'=cmd()"
        # Detalle is column 17 (index 16)
        assert row[17] == "'@SUM(1+1)*cmd()"

    def test_export_truncates_past_the_cap(self, admin_manager):
        for i in range(3):
            Pasaje.objects.create(
                tipo='B', dni=f'{i:08d}', nombres=f'Persona {i}',
                tipo_trabajador='WORKER', creado_por=admin_manager,
            )
        client = _client_for(admin_manager)

        with mock.patch('apps.core.utils.excel.MAX_EXPORT_ROWS', 1):
            response = client.get('/api/v1/administracion/pasajes/export/')

        assert response.status_code == status.HTTP_200_OK
        wb = _read_workbook(response)
        ws = wb.active
        warning_cell = ws.cell(row=1, column=1).value
        assert 'ADVERTENCIA' in warning_cell
