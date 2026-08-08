"""
API views for the Administración — Pasajes module.

All ViewSets require the user to have is_staff=True or the ADMIN_MANAGER role.
"""

import io
import logging
from datetime import date

import requests as http_requests

from django.conf import settings
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from django.core.cache import cache

from rest_framework import status, viewsets
from rest_framework.decorators import action, api_view, permission_classes as perm_classes
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import BasePermission, IsAuthenticated
from rest_framework.response import Response

from apps.administracion.models import (
    Pasaje, PoliticaPasajeDevoluciones, ProveedorPasajes, Puesto,
)
from apps.administracion.serializers import (
    PasajeCreateUpdateSerializer,
    PasajeListSerializer,
    PoliticaPasajeDevolucionesSerializer,
    ProveedorPasajesSerializer,
    PuestoSerializer,
)
from apps.administracion import services as adm_svc

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Permission
# ---------------------------------------------------------------------------

class IsAdminOrAdminManager(BasePermission):
    """
    Allow access to the Administración (Pasajes) module for:
      - Django staff/superusers (is_staff=True)
      - Users with the ADMIN_MANAGER or GENERAL_MANAGER role
      - Users with the dedicated PASAJES_MANAGER role
    """

    message = 'Se requiere rol Administrador, Gerente Administrativo o Gestor de Pasajes.'

    ALLOWED_ROLES = ['ADMIN_MANAGER', 'GENERAL_MANAGER', 'PASAJES_MANAGER']

    def has_permission(self, request, view):
        user = request.user
        if not user or not user.is_authenticated:
            return False
        if user.is_staff or user.is_superuser:
            return True
        return user.user_roles.filter(role__in=self.ALLOWED_ROLES).exists()


# ---------------------------------------------------------------------------
# Pagination
# ---------------------------------------------------------------------------

class AdmPagination(PageNumberPagination):
    page_size = 25
    page_size_query_param = 'page_size'
    max_page_size = 200


# ---------------------------------------------------------------------------
# Tipo de Cambio (SUNAT)
# ---------------------------------------------------------------------------

TIPO_CAMBIO_CACHE_KEY = 'sunat_tipo_cambio_hoy'

@api_view(['GET'])
@perm_classes([IsAuthenticated])
def tipo_cambio_hoy(request):
    """
    Returns today's SUNAT exchange rate (compra / venta).
    Caches the result for 4 hours to avoid hammering the external API.
    """
    cached = cache.get(TIPO_CAMBIO_CACHE_KEY)
    if cached:
        return Response(cached)

    try:
        resp = http_requests.get(
            'https://api.apis.net.pe/v1/tipo-cambio-sunat',
            timeout=10,
        )
        if resp.status_code == 200:
            data = resp.json()
            result = {
                'compra': float(data.get('compra', 0)),
                'venta': float(data.get('venta', 0)),
                'fecha': data.get('fecha', ''),
            }
            cache.set(TIPO_CAMBIO_CACHE_KEY, result, 60 * 60 * 4)
            return Response(result)
    except Exception:
        logger.warning('Error consultando tipo de cambio SUNAT', exc_info=True)

    return Response(
        {'compra': 0, 'venta': 0, 'fecha': ''},
        status=status.HTTP_503_SERVICE_UNAVAILABLE,
    )


# ---------------------------------------------------------------------------
# Puesto ViewSet
# ---------------------------------------------------------------------------

class PuestoViewSet(viewsets.ModelViewSet):
    """Lookup table for job positions / titles."""

    permission_classes = [IsAuthenticated, IsAdminOrAdminManager]
    serializer_class = PuestoSerializer
    queryset = Puesto.objects.filter(habilitado=True)
    pagination_class = None


# ---------------------------------------------------------------------------
# ProveedorPasajes ViewSet
# ---------------------------------------------------------------------------

class ProveedorPasajesViewSet(viewsets.ModelViewSet):
    """Full CRUD for travel ticket suppliers."""

    permission_classes = [IsAuthenticated, IsAdminOrAdminManager]
    serializer_class = ProveedorPasajesSerializer
    queryset = ProveedorPasajes.objects.all()
    pagination_class = AdmPagination

    def get_queryset(self):
        qs = ProveedorPasajes.objects.all()
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(ruc__icontains=search) | Q(razon_social__icontains=search)
            )
        return qs

    @action(detail=False, methods=['get'], url_path='buscar-ruc')
    def buscar_ruc(self, request):
        """
        Search for a supplier by exact RUC.
        If found locally, returns the stored record.
        If not found locally, queries the SUNAT API (apisperu.com),
        auto-creates the provider, and returns the result.
        """
        ruc = request.query_params.get('ruc', '').strip()
        if not ruc or len(ruc) != 11:
            return Response(
                {'detail': 'Se requiere un RUC de 11 dígitos.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # 1) Check local DB
        try:
            proveedor = ProveedorPasajes.objects.get(ruc=ruc)
            return Response({
                'found': True,
                'source': 'local',
                **ProveedorPasajesSerializer(proveedor).data,
            })
        except ProveedorPasajes.DoesNotExist:
            pass

        # 2) Query SUNAT API
        token = getattr(settings, 'RUC_API_TOKEN', '')
        api_url = getattr(settings, 'RUC_API_URL', '')
        if not token or not api_url:
            return Response(
                {'found': False, 'ruc': ruc, 'razon_social': ''},
                status=status.HTTP_200_OK,
            )

        try:
            resp = http_requests.get(
                f'{api_url}{ruc}',
                params={'token': token},
                timeout=10,
            )
            if resp.status_code == 200:
                data = resp.json()
                razon_social = (data.get('razonSocial') or '').strip()
                if razon_social:
                    proveedor = ProveedorPasajes.objects.create(
                        ruc=ruc,
                        razon_social=razon_social,
                    )
                    return Response({
                        'found': True,
                        'source': 'sunat',
                        **ProveedorPasajesSerializer(proveedor).data,
                    })
        except Exception:
            logger.warning('Error consultando RUC %s en SUNAT API', ruc, exc_info=True)

        return Response(
            {'found': False, 'ruc': ruc, 'razon_social': ''},
            status=status.HTTP_200_OK,
        )


# ---------------------------------------------------------------------------
# PoliticaPasajeDevoluciones ViewSet
# ---------------------------------------------------------------------------

class PoliticaPasajeDevolucionesViewSet(viewsets.ModelViewSet):
    """Full CRUD for refund policies per worker type and route."""

    permission_classes = [IsAuthenticated, IsAdminOrAdminManager]
    serializer_class = PoliticaPasajeDevolucionesSerializer
    queryset = PoliticaPasajeDevoluciones.objects.all()
    pagination_class = AdmPagination

    def get_queryset(self):
        qs = PoliticaPasajeDevoluciones.objects.all()
        habilitado = self.request.query_params.get('habilitado')
        if habilitado is not None:
            qs = qs.filter(habilitado=habilitado.lower() == 'true')
        return qs

    @action(detail=False, methods=['get'], url_path='por-tipo')
    def por_tipo(self, request):
        """Filter policies by tipo_trabajador (STAFF or WORKER)."""
        tipo = request.query_params.get('tipo', '').upper()
        if tipo not in ('STAFF', 'WORKER'):
            return Response(
                {'detail': "El parámetro 'tipo' debe ser STAFF o WORKER."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        qs = self.get_queryset().filter(tipo_trabajador=tipo, habilitado=True)
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)


# ---------------------------------------------------------------------------
# Pasaje ViewSet
# ---------------------------------------------------------------------------

class PasajeViewSet(viewsets.ModelViewSet):
    """
    Full CRUD for travel ticket records.

    Extra actions:
      - por_dni  — list all tickets for a given DNI
      - calcular_devolucion — compute refund based on policy
      - marcar_pagado  — mark a ticket as PAGADO
      - export  — download filtered list as Excel
    """

    permission_classes = [IsAuthenticated, IsAdminOrAdminManager]
    pagination_class = AdmPagination

    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return PasajeCreateUpdateSerializer
        return PasajeListSerializer

    def get_queryset(self):
        qs = Pasaje.objects.select_related(
            'proveedor', 'centro_costo', 'creado_por', 'actualizado_por'
        ).filter(habilitado=True)
        return self._apply_filters(qs, self.request.query_params)

    def _apply_filters(self, qs, params):
        """Shared filter helper — used in list() and export()."""
        dni = params.get('dni')
        tipo = params.get('tipo')
        estado = params.get('estado')
        mes = params.get('mes')
        fecha_from = params.get('fecha_from')
        fecha_to = params.get('fecha_to')
        search = params.get('search')
        tipo_trabajador = params.get('tipo_trabajador')
        centro_costo = params.get('centro_costo')
        moneda = params.get('moneda')

        if dni:
            qs = qs.filter(dni__icontains=dni)
        if tipo:
            qs = qs.filter(tipo=tipo)
        if estado:
            qs = qs.filter(estado=estado)
        if mes:
            qs = qs.filter(mes__iexact=mes)
        if fecha_from:
            qs = qs.filter(fecha__gte=fecha_from)
        if fecha_to:
            qs = qs.filter(fecha__lte=fecha_to)
        if tipo_trabajador:
            qs = qs.filter(tipo_trabajador=tipo_trabajador.upper())
        if centro_costo:
            qs = qs.filter(centro_costo_id=centro_costo)
        if moneda:
            qs = qs.filter(moneda=moneda.upper())
        if search:
            qs = qs.filter(
                Q(dni__icontains=search) |
                Q(nombres__icontains=search) |
                Q(factura_ticket__icontains=search) |
                Q(ruc__icontains=search) |
                Q(razon_social__icontains=search)
            )
        return qs

    # -- create: inject creado_por + auto-link personal ----------------------

    def perform_create(self, serializer):
        from apps.core.models import Personal
        personal_obj = None
        dni = serializer.validated_data.get('dni')
        if dni:
            personal_obj = Personal.objects.filter(dni=dni).first()
        serializer.save(creado_por=self.request.user, personal=personal_obj)

    # -- update: inject actualizado_por --------------------------------------

    def perform_update(self, serializer):
        serializer.save(actualizado_por=self.request.user)

    # -- por_dni ---------------------------------------------------------------

    @action(detail=False, methods=['get'], url_path='por-dni')
    def por_dni(self, request):
        """Return all active pasajes for a given DNI."""
        dni = request.query_params.get('dni', '').strip()
        if not dni:
            return Response(
                {'detail': "Se requiere el parámetro 'dni'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        pasajes = adm_svc.buscar_por_dni(dni)
        serializer = PasajeListSerializer(pasajes, many=True)
        return Response(serializer.data)

    # -- calcular_devolucion ---------------------------------------------------

    @action(detail=False, methods=['post'], url_path='calcular-devolucion')
    def calcular_devolucion(self, request):
        """
        Compute the devolution amount for a ticket against a specific policy.

        Request body:
          - monto         (Decimal, required)
          - moneda        ('SOLES' | 'DOLARES', required)
          - tipo_cambio   (Decimal, required when moneda=DOLARES)
          - politica_id   (int, required)

        Response:
          - devolucion    (Decimal, in PEN)
        """
        from decimal import Decimal, InvalidOperation

        monto_raw = request.data.get('monto')
        moneda = request.data.get('moneda', 'SOLES').upper()
        tipo_cambio_raw = request.data.get('tipo_cambio', '0')
        politica_id = request.data.get('politica_id')

        errors = {}
        if monto_raw is None:
            errors['monto'] = 'Campo requerido.'
        if not politica_id:
            errors['politica_id'] = 'Campo requerido.'
        if moneda not in ('SOLES', 'DOLARES'):
            errors['moneda'] = "Debe ser 'SOLES' o 'DOLARES'."
        if errors:
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            monto = Decimal(str(monto_raw))
        except InvalidOperation:
            return Response(
                {'monto': 'Valor numérico inválido.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            tipo_cambio = Decimal(str(tipo_cambio_raw))
        except InvalidOperation:
            tipo_cambio = Decimal('0')

        try:
            politica = PoliticaPasajeDevoluciones.objects.get(pk=politica_id)
        except PoliticaPasajeDevoluciones.DoesNotExist:
            return Response(
                {'politica_id': 'Política de devolución no encontrada.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        devolucion = adm_svc.calcular_devolucion(monto, moneda, tipo_cambio, politica)
        return Response({'devolucion': str(devolucion)})

    # -- marcar_pagado ---------------------------------------------------------

    @action(detail=True, methods=['post'], url_path='marcar-pagado')
    def marcar_pagado(self, request, pk=None):
        """
        Mark a ticket as PAGADO.

        Request body:
          - fecha_pago       (date string YYYY-MM-DD, optional — defaults to today)
          - numero_operacion (str, optional)
        """
        pasaje = self.get_object()

        if pasaje.estado == 'PAGADO':
            return Response(
                {'detail': 'El pasaje ya está marcado como pagado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        fecha_pago_raw = request.data.get('fecha_pago')
        numero_operacion = request.data.get('numero_operacion', '').strip()

        if fecha_pago_raw:
            try:
                from datetime import datetime
                fecha_pago = datetime.strptime(fecha_pago_raw, '%Y-%m-%d').date()
            except ValueError:
                return Response(
                    {'fecha_pago': 'Formato de fecha inválido. Use YYYY-MM-DD.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            fecha_pago = date.today()

        pasaje.estado = 'PAGADO'
        pasaje.fecha_pago = fecha_pago
        pasaje.numero_operacion = numero_operacion
        pasaje.actualizado_por = request.user
        pasaje.save(update_fields=[
            'estado', 'fecha_pago', 'numero_operacion',
            'actualizado_por', 'fecha_actualizacion',
        ])

        return Response(PasajeListSerializer(pasaje).data)

    # -- export ---------------------------------------------------------------

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        """Export the filtered pasajes list as an .xlsx file."""
        qs = self._apply_filters(
            Pasaje.objects.select_related(
                'proveedor', 'centro_costo', 'creado_por'
            ).filter(habilitado=True),
            request.query_params,
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Pasajes'

        headers = [
            'ID', 'Tipo', 'Fecha Bajada', 'Embarque Bajada', 'Destino Bajada',
            'Fecha Subida', 'Embarque Subida', 'Destino Subida',
            'DNI', 'Nombres', 'Cargo', 'Tipo Trabajador', 'Centro de Costo',
            'Proveedor', 'RUC', 'Razón Social', 'Factura/Ticket',
            'Detalle', 'Fecha', 'Mes',
            'Moneda', 'Monto IGV Soles', 'Monto IGV Dolares',
            'Tipo Cambio', 'Devolución', 'Total',
            'Estado', 'Fecha Pago', 'N. Operación',
            'Registrado por', 'Fecha Registro',
        ]
        ws.append(headers)

        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        bold_white = Font(bold=True, color='FFFFFF')
        for cell in ws[1]:
            cell.font = bold_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for p in qs:
            ws.append([
                p.id,
                p.get_tipo_display(),
                p.fecha_bajada.strftime('%d/%m/%Y') if p.fecha_bajada else '',
                p.embarque_bajada,
                p.destino_bajada,
                p.fecha_subida.strftime('%d/%m/%Y') if p.fecha_subida else '',
                p.embarque_subida,
                p.destino_subida,
                p.dni,
                p.nombres,
                p.cargo,
                p.get_tipo_trabajador_display(),
                p.centro_costo.name if p.centro_costo_id else '',
                p.proveedor.razon_social if p.proveedor_id else '',
                p.ruc,
                p.razon_social,
                p.factura_ticket,
                p.detalle,
                p.fecha.strftime('%d/%m/%Y') if p.fecha else '',
                p.mes,
                p.get_moneda_display(),
                float(p.monto_con_igv_soles),
                float(p.monto_con_igv_dolares),
                float(p.tipo_cambio),
                float(p.devolucion),
                float(p.total),
                p.get_estado_display(),
                p.fecha_pago.strftime('%d/%m/%Y') if p.fecha_pago else '',
                p.numero_operacion,
                p.creado_por.get_full_name() if p.creado_por_id else '',
                p.fecha_registro.strftime('%d/%m/%Y %H:%M') if p.fecha_registro else '',
            ])

        for col_idx, _ in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 18

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="pasajes.xlsx"'
        wb.save(response)
        return response
